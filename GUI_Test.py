import customtkinter as ctk
import math
import pandas as pd 
import os

from openpyxl import load_workbook
from datetime import datetime
# import matplotlib.pyplot as plt
# from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from Main_Code_2 import Calculate_Possible_Combinations
from Set_Default_Values import Set_Default_Values_For_GUI, Values_From_Boxes
from Entry_boxes_and_sliders import Make_Entry_boxes_and_sliders, Make_Sliders, Make_Sliders_desired_values, Make_Sliders_1_sf
from Compare_Best_Combinations import Compare_Best_Combination_changed_weightings
from Scatter_plots import plot_scatter, plot_scatter_current_options
from Create_parallel_coordinates_plot import create_parallel_coordinates_plot
from Get_battery_and_WLTP_data import Get_battery_and_WLTP_data
from Range_Estimation import Range_Estimation_for_Each_Battery
from Power_estimation_for_each_battery import Power_Estimation_for_Each_Battery


# Create the main window
app = ctk.CTk()
app.title("Tool to analyse how two battery chemistries can be combined to meet desired EV characteristics")
app.geometry("1350x800")

# frame = ctk.CTkFrame(app)
# frame.grid(row=0, column=0, padx=20, pady=20)

global has_calculate_been_pressed, car_data
has_calculate_been_pressed = 0

# Nissan Leaf
input_range = 170
input_energy = 28000
input_discharging_power = 90000
input_max_V = 400
input_min_V = 240
input_max_mass_battery = 185.5
input_max_mass_pack = 315
input_charging_power = 50000
input_max_volume = 0.485

# car_data = [3100, 0.3, 3.38, 0.015, 0] # Rivian R1T             Actual: 505, Calculated: 508
# car_data = [1748, 0.29, 2.37, 0.015, 0] # Kia Niro EV         Actual: 384, Calculated: 405
car_data = [1486, 0.28, 2.32, 0.015, 0] # Nissan Leaf         Actual: 169(excel)/135, Calculated: 179 Using 2 chems: 310
# car_data = [1830, 0.23, 2.268, 0.015, 0] # Tesla model 3      Actual: 576, Calculated: 572
# car_data = [2584, 0.29, 2.3, 0.015, 0] # Polestar 3              Actual: 482, Calculated: 532

# # Normal input values
# input_range = 500
# input_energy = 75000
# input_discharging_power = 250000
# input_max_V = 450
# input_min_V = 280
# input_max_mass_battery = 320
# input_max_mass_pack = 500
# input_charging_power = 150000

# # Initial car data
# car_data[2000, 0.29, 2.35, 0.015]

# Max_range_row = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
# min_charging_time_row = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]

try:
    for successful_combinations, best_weighted_normaliesed, count_successful_combinations, count_successful_combinations_1_bat, total_checked in \
    Calculate_Possible_Combinations(input_energy, input_discharging_power, input_max_V, input_min_V, input_max_mass_pack, input_charging_power, car_data, input_max_volume):
        if total_checked == 10:
            print(f"Total checked: {total_checked}")

except StopIteration as e:
    successful_combinations, best_weighted_normaliesed, count_successful_combinations, count_successful_combinations_1_bat, total_checked = e.value
print(f"Battery 2 combinations: {count_successful_combinations}, Battery 1 combinations: {count_successful_combinations_1_bat}, Total checked: {total_checked}")

successful_combinations_1_bat = None
successful_combinations_2_bat = None

# Function to calculate based on slider and input values
def calculate():
    calculate_button_label.configure(text=f"Calculating...\n \n")
    result_label_weightings.configure(text=f" \n \n \n \n \n ")
    result_label_desired_values.configure(text=f"")
    

    global has_calculate_been_pressed, car_data
    has_calculate_been_pressed = 1

    app.update()
    

    print(car_data[2], car_data)

    # [float(EV_range.get()), float(total_energy.get()), float(Pack_mass.get()), float(Max_V.get()), float(Min_V.get()), float(Discharging_power.get()), float(Charging_power.get())]

    # desired_EV_characteristics = [EV_range.get(), total_energy.get(), Pack_mass.get(), Max_V.get(), Min_V.get(), Discharging_power.get(), Charging_power.get()]
    # slider_values = [float(EV_range_slider.get()), float(total_energy_slider.get()), float(Pack_mass_slider.get()), float(Max_V_slider.get()), float(Min_V_slider.get()),float(Discharging_power_slider.get()), float(Charging_power_slider.get())]
    desired_EV_characteristics = [0, total_energy.get(), Pack_mass.get(), Max_V.get(), Min_V.get(), Discharging_power.get(), Charging_power.get(), Volume.get()]
    slider_values = [0, float(total_energy_slider.get()), float(Pack_mass_slider.get()), float(Max_V_slider.get()), float(Min_V_slider.get()),float(Discharging_power_slider.get()), float(Charging_power_slider.get()), float(Volume_slider.get())]
    
    car_data = [EV_mass.get(), EV_drag.get(), EV_front_area.get(), EV_r_r.get()]
    EV_slider_values = [float(EV_mass_slider.get()), float(EV_drag_slider.get()), round(float(EV_front_area_slider.get()), 4), float(EV_r_r_slider.get())]
    # slider_values = [range_slider.get(), total_energy_slider.get(), Pack_mass_slider.get(), Max_V_slider.get(), Min_V_slider.get(), Discharging_power_slider.get(), Charging_power_slider.get()]


    for i in range(len(desired_EV_characteristics)):
        # print(desired_EV_characteristics[i])

        if desired_EV_characteristics[i] == "":
            desired_EV_characteristics[i] = slider_values[i]
        # elif i == 0: 
        #     try: desired_EV_characteristics[i] = float(EV_range.get())
        #     except: print("non number entered in EV")
        elif i == 1: 
            try: desired_EV_characteristics[i] = float(total_energy.get())
            except: print("non number entered in energy")
        elif i == 2: 
            try: desired_EV_characteristics[i] = float(Pack_mass.get())
            except: print("non number entered in pack mass")
        elif i == 3: 
            try: desired_EV_characteristics[i] = float(Max_V.get())
            except: print("non number entered in max voltage")
        elif i == 4: 
            try: desired_EV_characteristics[i] = float(Min_V.get())
            except: print("non number entered in min voltage")
        elif i == 5: 
            try: desired_EV_characteristics[i] = float(Discharging_power.get())
            except: print("non number entered in discharging")
        elif i == 6: 
            try: desired_EV_characteristics[i] = float(Charging_power.get())
            except: print("non number entered in charging")
        elif i == 7: 
            try: desired_EV_characteristics[i] = float(Volume.get())
            except: print("non number entered in volume")
        
    # if desired_EV_characteristics[5] > desired_EV_characteristics[6]:
    #     # print(desired_EV_characteristics[2])
    #     desired_EV_characteristics[2] = desired_EV_characteristics[2] - (math.ceil(desired_EV_characteristics[5]/48200))
    #     # print(desired_EV_characteristics[2])
    #     desired_EV_characteristics[7] = desired_EV_characteristics[7] - (math.ceil(desired_EV_characteristics[5]/48200)) * 0.00018333
    # else:
    #     desired_EV_characteristics[2] = desired_EV_characteristics[2] - (math.ceil(desired_EV_characteristics[6]/48200))
    #     desired_EV_characteristics[7] = desired_EV_characteristics[7] - (math.ceil(desired_EV_characteristics[5]/48200)) * 0.00018333

    for i in range(len(car_data)):
        if car_data[i] == "":
            car_data[i] = EV_slider_values[i]
        elif i == 0: 
            try: car_data[i] = float(EV_mass.get())
            except: print("non number entered in EV mass")
        elif i == 1: 
            try: car_data[i] = float(EV_drag.get())
            except: print("non number entered in EV drag")
        elif i == 2: 
            try: car_data[i] = float(EV_front_area.get())
            except: print("non number entered in EV frontal area")
        elif i == 3: 
            try: car_data[i] = float(EV_r_r.get())
            except: 
                print("non number entered in EV rolling resistance")

    req_range = desired_EV_characteristics[0]
    req_energy = desired_EV_characteristics[1]
    req_max_mass_pack = desired_EV_characteristics[2]
    req_max_V = desired_EV_characteristics[3]
    req_min_V = desired_EV_characteristics[4]
    req_discharging_power = desired_EV_characteristics[5]
    req_charging_power = desired_EV_characteristics[6]
    max_volume = desired_EV_characteristics[7]


    # req_range, req_energy, req_discharging_power, req_max_V, req_min_V, req_max_mass_pack, req_charging_power = Values_From_Boxes(float(range.get()), float(total_energy.get()), float(Discharging_power.get()), float(Charging_power.get()), float(Max_V.get()), float(Min_V.get()), float(Pack_mass.get()), req_range, req_energy, req_discharging_power, req_max_V, req_min_V, req_max_mass_pack, req_charging_power)
    print(f"Reqired values: {desired_EV_characteristics}")
    print(f"EV values{car_data}")
    
    global successful_combinations, successful_combinations_both, count_successful_combinations_2_bat, count_successful_combinations_1_bat, successful_combinations_1_bat, successful_combinations_2_bat

    # for successful_combinations, best_weighted_normaliesed, count_successful_combinations_2_bat, count_successful_combinations_1_bat, total_checked in \
    #     Calculate_Possible_Combinations(req_energy, req_discharging_power, req_max_V, req_min_V, req_max_mass_pack, req_charging_power, EV_metrics):

    #     result_label_calculating.configure(text=f"Calculating...\nBattery Combinations Found: {count_successful_combinations_2_bat}\nTotal Checked: {total_checked}")
    #     app.update()

    try:
        for successful_combinations, best_weighted_normaliesed, count_successful_combinations_2_bat, count_successful_combinations_1_bat, total_checked in \
        Calculate_Possible_Combinations(req_energy, req_discharging_power, req_max_V, req_min_V, req_max_mass_pack, req_charging_power, car_data, max_volume):
            if total_checked % 1000 == 0:
                calculate_button_label.configure(text=f"Calculating...\nBattery Combinations Found: {count_successful_combinations_2_bat}\nTotal Checked: {total_checked}")
                app.update()

    except StopIteration as e:
        successful_combinations, best_weighted_normaliesed, count_successful_combinations_2_bat, count_successful_combinations_1_bat, total_checked = e.value

    if count_successful_combinations_1_bat == 0:
        print("No 1 battery combinations found")
        successful_combinations_1_bat = 0
        successful_combinations_2_bat = successful_combinations
        successful_combinations_both = successful_combinations
        
    elif count_successful_combinations_2_bat == 0:
        print("No 2 battery combinations found")

        successful_combinations_1_bat = successful_combinations
        successful_combinations_2_bat = 0
        successful_combinations_both = successful_combinations

    else:
        successful_combinations_1_bat = successful_combinations[-count_successful_combinations_1_bat:]
        successful_combinations_2_bat = successful_combinations[:-count_successful_combinations_1_bat]
        successful_combinations_both = successful_combinations

    if len(successful_combinations) > 0:
        options = len(successful_combinations)
        clear_column(7)
        update_numbers_for_desired_sliders()
        make_desired_sliders() # Make the desired sliders using the values calculated for the max and middlie values
        Make_weightings_sliders()
        reset_weightings_button.grid(row= 10, column= 6, padx=0, pady=0)
        
        # checkbox_Only_2_batteries.grid(row=13, column=7, padx=(5, 5), pady=(0, 0), sticky="w")
        # checkbox_Only_1_battery.grid(row=14, column=7, padx=(5, 5), pady=(20, 0), sticky="w")
        # checkbox_both_batteries.grid(row=15, column=7, padx=(5, 5), pady=(20, 0), sticky="w")
        checkbox_Only_2_batteries.grid(row=14, column=7, padx=(5, 5), pady=(0, 0), sticky="w")
        checkbox_Only_1_battery.grid(row=15, column=7, padx=(5, 5), pady=(0, 0), sticky="w")
        checkbox_both_batteries.grid(row=16, column=7, padx=(5, 5), pady=(0, 0), sticky="w")
        plot_button.grid(row=14, column=8, pady=0, padx=0)
        # plot_button_current_options.grid(row=15, column=8, pady=10, padx=0)
        parallel_coordinates_plot_button.grid(row=15, column=8, pady=5, padx=0)
        # excel_output_button.grid(row= 9, column= 8, padx=0, pady=0)
        excel_output_all_button.grid(row= 10, column= 8, padx=0, pady=0)
        BMS_option_button.grid(row= 11, column= 8, padx=0, pady=0)
        vertical_line.place(x=832, y=0)

        # result_label_desired_values.configure(text=f"Adjust desired targets to see result")
        desired_EV_outputs_title_label.configure(text="Please adjust sliders for desired\nEV outputs")
        Weightings_sliders_title_label.configure(text="Please adjust sliders for desired\nweightings")

        # Set the weighting and desired sliders to start working
        Range_weighting_slider.configure(state="normal")
        Min_charging_time_weighting_slider.configure(state="normal")
        max_discharge_power_weighting_slider.configure(state="normal")
        min_pack_mass_weighting_slider.configure(state="normal")
        Desired_range_slider.configure(state="normal")
        Desired_km_per_min_slider.configure(state="normal")
        # Desired_min_charging_time_slider.configure(state="normal")
        Desired_max_discharge_power_slider.configure(state="normal")
        Desired_max_mass_slider.configure(state="normal")
        # min_charging_time_slider.configure(state="normal")

        
        
        result_label_weightings.configure(text=f"Optimal weightings result:\nRange: {best_weighted_normaliesed[10]:.2f}(km)\nCharging speed: {((0.8*best_weighted_normaliesed[10] - 0.1*best_weighted_normaliesed[10])/best_weighted_normaliesed[11]):.0f}(km/min) \n Max discharging power: {(best_weighted_normaliesed[7]/1000):.0f}(kW) \n Min pack mass: {best_weighted_normaliesed[8]:.0f}(kg) \n Max charging power: {(best_weighted_normaliesed[9]/1000):.0f}(kW)")
        calculate_button_label.configure(text=f"Options: {options}\n 2 battery combinations: {count_successful_combinations_2_bat}\n1 battery combinations: {count_successful_combinations_1_bat}")
    else:
        calculate_button_label.configure(text=f"No combinations found")
        # result_label_weightings.configure(text=f"Press Calculate to see results \n \n \n \n \n ")
        # result_label_desired_values.configure(text=f"Press Calculate to see results \n \n \n \n \n ")
        app.update()
    
    # return desired_EV_characteristics, car_data


    

# def clear_column(column):
#     for widget in app.winfo_children():
#         if widget.grid_info().get("column") == column:
#             widget.grid_remove()

def clear_column(column):
    for widget in app.winfo_children():
        if isinstance(widget, ctk.CTkToplevel):
            continue
        grid_info = widget.grid_info()
        if grid_info.get("column") == column and 1 <= grid_info.get("row") <= 8:
            widget.grid_remove()
                

def check_selected(value):
    checkbox_Only_1_battery.deselect()
    checkbox_Only_2_batteries.deselect()
    checkbox_both_batteries.deselect()

    print(value)

    if value == 2:
        checkbox_Only_1_battery.select()
        if successful_combinations_1_bat != 0:
            global successful_combinations
            successful_combinations = successful_combinations_1_bat
            on_desired_slider_release()
        else:
            print("No 1 battery combinations found")
            result_label_desired_values.configure(text=f"No 1 Battery Options")

    elif value == 1:
        checkbox_Only_2_batteries.select()
        if successful_combinations_2_bat != 0:
            Only_2_bat()
        else:
            print("No 2 battery combinations found")
            result_label_desired_values.configure(text=f"No 2 Battery Options")
          
    elif value == 3:
        checkbox_both_batteries.select()
        if successful_combinations_2_bat != 0:
            Both_bat()
        else:
            print("No battery combinations found")
            result_label_desired_values.configure(text=f"No Battery Options")
       
        
def Only_2_bat():
    global successful_combinations
    successful_combinations = successful_combinations_2_bat
    on_desired_slider_release()

def Both_bat():
    global successful_combinations
    successful_combinations = successful_combinations_both
    on_desired_slider_release()


def non_number_message():
    # Create the label
    message_label = ctk.CTkLabel(app, text="Please enter a valid number", font=(None, 25), fg_color="red", text_color="black", 
                                 corner_radius=5, width=220, height=28)
    message_label.grid(row=9, column=3, columnspan=2, padx=10, pady=10)
    
    # Remove the label after 5 seconds
    app.after(2000, message_label.destroy)

def update_total_energy_label(value):
    total_energy_label.configure(text=f"Energy (Wh): {float(value):.0f}")

def update_Pack_mass_label(value):
    Pack_mass_label.configure(text=f"Pack Mass (kg): {float(value):.0f}")

def update_Max_V_label(value):
    Max_V_label.configure(text=f"Maximum Voltage (V): {float(value):.0f}")

def update_Min_V_label(value):
    Min_V_label.configure(text=f"Minimum Voltage (V): {float(value):.0f}")

def update_Discharging_power_label(value):
    Discharging_power_label.configure(text=f"Discharging Power (W): {float(value):.0f}")

def update_Charging_power_label(value):
    Charging_power_label.configure(text=f"Charging Power (W): {float(value):.0f}")

def update_Volume_label(value):
    Volume_label.configure(text=f"Pack Volume (m³): {float(value):.3f}")

def update_EV_mass_label(value):
    EV_mass_label.configure(text=f"EV Mass Without Pack (kg): {float(value):.0f}")

def update_EV_drag_label(value):
    EV_drag_label.configure(text=f"EV Drag Coefficient: {float(value):.2f}")

def update_EV_front_area_label(value):
    EV_front_area_label.configure(text=f"EV Frontal Area (m²): {float(value):.2f}")

def update_EV_r_r_label(value):
    EV_r_r_label.configure(text=f"EV Rolling Resistance (N): {float(value):.3f}")

# Weighting sliders
default_weighting = [1.15, 0.9, 0, 0, 0]

global Range_weighting, min_charging_time_weighting, max_discharge_power_weighting, min_pack_mass_weighting, min_charge_power_weighting

Range_weighting = default_weighting[0]
min_charging_time_weighting = default_weighting[1]
max_discharge_power_weighting = default_weighting[2]
min_pack_mass_weighting = default_weighting[3]
min_charge_power_weighting = default_weighting[4]

def update_range_weighting_label(value):
    Range_weighting_label.configure(text=f"Range Weighting: {float(value):.2f}")
    global Range_weighting
    Range_weighting = value
    
def update_min_charging_time_weighting_label(value):
    Min_charging_time_weighting_label.configure(text=f"Min Charging Time Weighting: {float(value):.2f}")
    global min_charging_time_weighting
    min_charging_time_weighting = value

def update_max_discharge_power_weighting_label(value):
    max_discharge_power_weighting_label.configure(text=f"Max Discharge Power Weighting: {float(value):.2f}")
    global max_discharge_power_weighting
    max_discharge_power_weighting = value

def update_min_pack_mass_weighting_label(value):
    min_pack_mass_weighting_label.configure(text=f"Min Pack Mass Weighting: {float(value):.2f}")
    global min_pack_mass_weighting
    min_pack_mass_weighting = value

# def update_min_charge_power_weighting_label(value):
#     min_charge_power_weighting_label.configure(text=f"Min Charge Power Weighting: {float(value):.2f}")
#     global min_charge_power_weighting
#     min_charge_power_weighting = value



def update_numbers_for_desired_sliders():
    global min_range_row, max_range_row, min_charging_time_row, max_charging_time_row, max_discharging_power_row, min_discharging_power_row, max_mass_row, min_mass_row, list_km_per_min
    
    max_range_row = max(successful_combinations, key=lambda x: x[10])
    min_range_row = min(successful_combinations, key=lambda x: x[10])
    min_charging_time_row = min(successful_combinations, key=lambda x: x[11])
    max_charging_time_row = max(successful_combinations, key=lambda x: x[11])
    max_discharging_power_row = max(successful_combinations, key=lambda x: x[7])
    min_discharging_power_row = min(successful_combinations, key=lambda x: x[7])
    min_mass_row = min(successful_combinations, key=lambda x: x[8])
    max_mass_row = max(successful_combinations, key=lambda x: x[8])

    list_km_per_min = []
    for i in range(0, len(successful_combinations)):
        km_per_min = (0.8*successful_combinations[i][10] - 0.1*successful_combinations[i][10])/successful_combinations[i][11]
        list_km_per_min.append(km_per_min)
    # print(f"km_per_min: {list_km_per_min}")


    global desired_range, desired_min_charging_time, desired_max_discharging_power, desired_max_mass, desired_km_per_min
    desired_range = min_range_row[10]
    desired_min_charging_time = max_charging_time_row[11]
    desired_max_discharging_power = min_discharging_power_row[7]/1000
    desired_max_mass = max_mass_row[8]
    desired_km_per_min = min(list_km_per_min)
    print(desired_km_per_min)



def update_desired_range_label(value):
    Desired_range_label.configure(text=f"Range (km): {float(value):.0f}")
    global desired_range
    desired_range = value

# def update_desired_min_charging_time_label(value):
#     Desired_min_charging_time_label.configure(text=f"Charging Time (10-80%) (mins): {float(value):.0f}")
#     global desired_min_charging_time
#     desired_min_charging_time = value

def update_desired_km_per_min_label(value):
    Desired_km_per_min_label.configure(text=f"Charging speed (km/min): {float(value):.1f}")
    global desired_km_per_min
    desired_km_per_min = value

def update_desired_max_discharge_power_label(value):
    Desired_max_discharge_power_label.configure(text=f"Max Discharge Power (kW): {float(value):.0f}")
    global desired_max_discharging_power
    desired_max_discharging_power = value

def update_desired_max_mass_label(value):
    Desired_max_mass_label.configure(text=f"Max Mass (kg): {float(value):.0f}")
    global desired_max_mass
    desired_max_mass = value

def update_sliders(event=None):
    try:
    
        value_total_energy_empty = total_energy.get()
        if value_total_energy_empty == "":
            total_energy_slider.set(input_energy)
            update_total_energy_label(input_energy)
        else:
            try: 
                value_total_energy = float(total_energy.get())
                if total_energy_slider._from_ <= value_total_energy <= total_energy_slider._to:
                    total_energy_slider.set(value_total_energy) 
                    update_total_energy_label(value_total_energy)
                else:
                    update_total_energy_label(value_total_energy)
            except: 
                print("non number entered in energy")
                non_number_message()

        value_Pack_mass_empty = Pack_mass.get()
        if value_Pack_mass_empty == "":
            Pack_mass_slider.set(input_max_mass_pack)
            update_Pack_mass_label(input_max_mass_pack)
        else:
            try:
                value_Pack_mass = float(Pack_mass.get())
                if Pack_mass_slider._from_ <= value_Pack_mass <= Pack_mass_slider._to:
                    Pack_mass_slider.set(value_Pack_mass) 
                    update_Pack_mass_label(value_Pack_mass)

                else:
                    update_Pack_mass_label(value_Pack_mass)
            except: 
                print("non number entered in mass")
                non_number_message()

        value_Max_V_empty = Max_V.get()
        if value_Max_V_empty == "":
            Max_V_slider.set(input_max_V)  # Assuming input_Max_V is a predefined default value
            update_Max_V_label(input_max_V)
        else:
            try:
                value_Max_V = float(Max_V.get())
                if Max_V_slider._from_ <= value_Max_V <= Max_V_slider._to:
                    Max_V_slider.set(value_Max_V)
                    update_Max_V_label(value_Max_V)
                else:
                    update_Max_V_label(value_Max_V)
            except:
                print("non number entered in Max_V")
                non_number_message()

        value_Min_V_empty = Min_V.get()
        if value_Min_V_empty == "":
            Min_V_slider.set(input_min_V)  # Assuming input_Min_V is a predefined default value
            update_Min_V_label(input_min_V)
        else:
            try:
                value_Min_V = float(Min_V.get())
                if Min_V_slider._from_ <= value_Min_V <= Min_V_slider._to:
                    Min_V_slider.set(value_Min_V)
                    update_Min_V_label(value_Min_V)
                else:
                    update_Min_V_label(value_Min_V)
            except:
                print("non number entered in Min_V")
                non_number_message()

        value_Discharging_power_empty = Discharging_power.get()
        if value_Discharging_power_empty == "":
            Discharging_power_slider.set(input_discharging_power)  # Assuming input_Discharging_power is a predefined default value
            update_Discharging_power_label(input_discharging_power)
        else:
            try:
                value_Discharging_power = float(Discharging_power.get())
                if Discharging_power_slider._from_ <= value_Discharging_power <= Discharging_power_slider._to:
                    Discharging_power_slider.set(value_Discharging_power)
                    update_Discharging_power_label(value_Discharging_power)
                else:
                    update_Discharging_power_label(value_Discharging_power)
            except:
                print("non number entered in Discharging_power")
                non_number_message()

        value_Charging_power_empty = Charging_power.get()
        if value_Charging_power_empty == "":
            Charging_power_slider.set(input_charging_power)  # Assuming input_Charging_power is a predefined default value
            update_Charging_power_label(input_charging_power)
        else:
            try:
                value_Charging_power = float(Charging_power.get())
                if Charging_power_slider._from_ <= value_Charging_power <= Charging_power_slider._to:
                    Charging_power_slider.set(value_Charging_power)
                    update_Charging_power_label(value_Charging_power)
                else:
                    update_Charging_power_label(value_Charging_power)
            except:
                print("non number entered in Charging_power")
                non_number_message()
        
        value_volume_empty = Volume.get()
        if value_volume_empty == "":
            Volume_slider.set(input_max_volume)  # Assuming input_Charging_power is a predefined default value
            update_Volume_label(input_max_volume)
        else:
            try:
                value_Volume = float(Volume.get())
                if Volume_slider._from_ <= value_Volume <= Volume_slider._to:
                    Volume_slider.set(value_Volume)
                    update_Volume_label(value_Volume)
                else:
                    update_Volume_label(value_Volume)
            except:
                print("non number entered in Volume")
                non_number_message()

        value_EV_mass = EV_mass.get()
        if value_EV_mass == "":
            EV_mass_slider.set(car_data[0])
            update_EV_mass_label(car_data[0])
        else:
            try: 
                value_EV_mass = float(EV_mass.get())  # Get the value from the entry box
                if EV_mass_slider._from_ <= value_EV_mass <= EV_mass_slider._to:  # Check if the value is within the slider's range
                    EV_mass_slider.set(value_EV_mass)  # Set the slider to the value
                    update_EV_mass_label(value_EV_mass)
                else:
                    update_EV_mass_label(value_EV_mass)

            except: 
                print("non number entered in EV mass")
                non_number_message()
        
        value_EV_drag = EV_drag.get()
        if value_EV_drag == "":
            EV_drag_slider.set(car_data[1])
            update_EV_drag_label(car_data[1])
        else:
            try: 
                value_EV_drag = float(EV_drag.get())  # Get the value from the entry box
                if EV_drag_slider._from_ <= value_EV_drag <= EV_drag_slider._to:  # Check if the value is within the slider's range
                    EV_drag_slider.set(value_EV_drag)  # Set the slider to the value
                    update_EV_drag_label(value_EV_drag)
                else:
                    update_EV_drag_label(value_EV_drag)

            except: 
                print("non number entered in EV drag")
                non_number_message()

        value_EV_front_area = EV_front_area.get()
        if value_EV_front_area == "":
            EV_front_area_slider.set(car_data[2])
            update_EV_front_area_label(car_data[2])
        else:
            try: 
                value_EV_front_area = float(EV_front_area.get())  # Get the value from the entry box
                if EV_front_area_slider._from_ <= value_EV_front_area <= EV_front_area_slider._to:  # Check if the value is within the slider's range
                    EV_front_area_slider.set(value_EV_front_area)  # Set the slider to the value
                    update_EV_front_area_label(value_EV_front_area)
                else:
                    update_EV_front_area_label(value_EV_front_area) 

            except: 
                print("non number entered in EV frontal area")
                non_number_message()

        value_EV_r_r = EV_r_r.get()
        if value_EV_r_r == "":
            EV_r_r_slider.set(car_data[3])
            update_EV_r_r_label(car_data[3])
        else:
            try: 
                value_EV_r_r = float(EV_r_r.get())  # Get the value from the entry box
                if EV_r_r_slider._from_ <= value_EV_r_r <= EV_r_r_slider._to:  # Check if the value is within the slider's range
                    EV_r_r_slider.set(value_EV_r_r)  # Set the slider to the value
                    update_EV_r_r_label(value_EV_r_r)
                else:
                    update_EV_r_r_label(value_EV_r_r)

            except: 
                print("non number entered in EV rolling resistance")
                non_number_message()

    except ValueError:
        print("Invalid input 2. Please enter a number.")

def on_weighting_slider_release(event=None):
    weighting = [Range_weighting, min_charging_time_weighting, max_discharge_power_weighting, min_pack_mass_weighting, min_charge_power_weighting]
    best_weighted_normaliesed = Compare_Best_Combination_changed_weightings(successful_combinations, weighting)
    if has_calculate_been_pressed == 1:
        result_label_weightings.configure(text=f"Optimal weightings result:\nRange: {best_weighted_normaliesed[10]:.2f}(km) \n Charging speed: {((0.8*best_weighted_normaliesed[10] - 0.1*best_weighted_normaliesed[10])/best_weighted_normaliesed[11]):.0f}(km/min) \n Max discharging power: {(best_weighted_normaliesed[7]/1000):.0f}(kW) \n Min pack mass: {best_weighted_normaliesed[8]:.0f}(kg) \n Max charging power: {(best_weighted_normaliesed[9]/1000):.0f}(kW)")
    else:
        result_label_weightings.configure(text="Press Calculate to see results \n \n \n \n \n ")

def reset_weightings():
    Range_weighting_slider.set(default_weighting[0])
    Min_charging_time_weighting_slider.set(default_weighting[1])
    max_discharge_power_weighting_slider.set(default_weighting[2])
    min_pack_mass_weighting_slider.set(default_weighting[3])
    # min_charge_power_weighting_slider.set(default_weighting[4])

    Range_weighting_label.configure(text=f"Range Weighting: {default_weighting[0]}")
    Min_charging_time_weighting_label.configure(text=f"Max Charging Speed Weighting: {default_weighting[1]}")
    max_discharge_power_weighting_label.configure(text=f"Max Discharge Power Weighting: {default_weighting[2]}")
    min_pack_mass_weighting_label.configure(text=f"Min Pack Mass Weighting: {default_weighting[3]}")
    # min_charge_power_weighting_label.configure(text=f"Min Charge Power Weighting: {default_weighting[4]}")

    update_range_weighting_label(default_weighting[0])
    update_min_charging_time_weighting_label(default_weighting[1])
    update_max_discharge_power_weighting_label(default_weighting[2])
    update_min_pack_mass_weighting_label(default_weighting[3])
    # update_min_charge_power_weighting_label(default_weighting[4])

    on_weighting_slider_release()

def reset_inputs():
    EV_mass_slider.set(car_data[0])
    EV_drag_slider.set(car_data[1])
    EV_front_area_slider.set(car_data[2])
    EV_r_r_slider.set(car_data[3])

    total_energy_slider.set(input_energy)
    Pack_mass_slider.set(input_max_mass_pack)
    Max_V_slider.set(input_max_V)
    Min_V_slider.set(input_min_V)
    Discharging_power_slider.set(input_discharging_power)
    Charging_power_slider.set(input_charging_power)
    Volume_slider.set(input_max_volume)

    total_energy.delete(0, 'end')
    total_energy.configure(placeholder_text="Total Energy (Wh)")
    Pack_mass.delete(0, 'end')
    Pack_mass.configure(placeholder_text="Pack Mass (kg)")
    Max_V.delete(0, 'end') 
    Max_V.configure(placeholder_text="Maximum Voltage (V)")
    Min_V.delete(0, 'end')
    Min_V.configure(placeholder_text="Minimum Voltage (V)")
    Discharging_power.delete(0, 'end')
    Discharging_power.configure(placeholder_text="Peak Discharging Power (W)")
    Charging_power.delete(0, 'end')
    Charging_power.configure(placeholder_text="Peak Charging Power (W)")
    EV_mass.delete(0, 'end')
    EV_mass.configure(placeholder_text="EV Mass Without Pack (kg)")
    EV_drag.delete(0, 'end')
    EV_drag.configure(placeholder_text="EV Drag Coefficient")
    EV_front_area.delete(0, 'end')
    EV_front_area.configure(placeholder_text="EV Frontal Area (m²)")
    EV_r_r.delete(0, 'end')
    EV_r_r.configure(placeholder_text="EV Rolling Resistance (N)")

    update_EV_mass_label(car_data[0])
    update_EV_drag_label(car_data[1])
    update_EV_front_area_label(car_data[2])
    update_EV_r_r_label(car_data[3])

    update_total_energy_label(input_energy)
    update_Pack_mass_label(input_max_mass_pack)
    update_Max_V_label(input_max_V)
    update_Min_V_label(input_min_V)
    update_Discharging_power_label(input_discharging_power)
    update_Charging_power_label(input_charging_power)

def on_desired_slider_release(event=None):
    desired_values = [desired_range, desired_km_per_min, desired_max_discharging_power, desired_max_mass]
    print(desired_values, has_calculate_been_pressed)
    if has_calculate_been_pressed == 1:
        matching_rows = [row for row in successful_combinations if row[10] >= desired_values[0] and (0.8*row[10] - 0.1*row[10])/row[11] >= desired_values[1]\
                         and row[7]/1000 >= desired_max_discharging_power and row[8] <= desired_max_mass]
        if matching_rows:
            
            # print(f"Matching rows length: {len(matching_rows)}")
            # print(matching_rows[0], matching_rows[len(matching_rows)-2], matching_rows[len(matching_rows)-1])

            result_label_desired_values.configure(text=f"Current Options: {len(matching_rows)} \nRange: {matching_rows[0][10]:.2f}(km) \nCharging speed: {((0.8*matching_rows[0][10] - 0.1*matching_rows[0][10])/matching_rows[0][11]):.1f}(km/min)\nMax discharging power: {(matching_rows[0][7]/1000):.0f}(kW) \nMin pack mass: {matching_rows[0][8]:.0f}(kg) \nMax charging power: {(matching_rows[0][9]/1000):.0f}(kW)")
        else:
            result_label_desired_values.configure(text=f"No matching rows found")
    else:
        result_label_desired_values.configure(text="Press Calculate to see results")


    # result_label.configure(text=f"Range: {best_weighted_normaliesed[10]:.2f}(km) \n Charging time(10-80%): {best_weighted_normaliesed[11]:.0f}(mins) \n Max discharging power: {(best_weighted_normaliesed[7]/1000):.0f}(kW) \n Min pack mass: {best_weighted_normaliesed[8]:.0f}(kg) \n Max charging power: {(best_weighted_normaliesed[9]/1000):.0f}(kW)")

max_range_row = max(successful_combinations, key=lambda x: x[10])
min_charging_time_row = min(successful_combinations, key=lambda x: x[11])
max_charging_time_row = max(successful_combinations, key=lambda x: x[11])
max_discharging_power_row = max(successful_combinations, key=lambda x: x[7])
Min_mass_row = min(successful_combinations, key=lambda x: x[8])
# desired_km_per_min = min(list_km_per_min)

list_km_per_min = []
for i in range(0, len(successful_combinations)):
    km_per_min = (0.8*successful_combinations[i][10] - 0.1*successful_combinations[i][10])/successful_combinations[i][11]
    list_km_per_min.append(km_per_min)
# print(f"km_per_min: {list_km_per_min}")


def make_desired_sliders():
    global Desired_range_slider, Desired_range_label, Desired_km_per_min_slider, Desired_km_per_min_label, Desired_max_discharge_power_slider, Desired_max_discharge_power_label, Desired_max_mass_slider, Desired_max_mass_label, result_label_desired_values,Desired_min_charging_time_slider, Desired_min_charging_time_label, min_charging_time_slider

    Desired_range_slider = ctk.CTkSlider(app, from_= 0, to=round(math.floor(max_range_row[10]), -1), number_of_steps=round(math.floor(max_range_row[10]), -1)/20)
    Desired_range_slider.grid(row= 1, column=7, padx=10, pady=10)
    Desired_range_slider.set(max_range_row[10])
    # total_energy_slider.configure(command=update_total_energy_label)
    Desired_range_label = ctk.CTkLabel(app, text=default_weighting[0])
    Desired_range_label.grid(row= 2, column=7)

    # Desired_range_slider, Desired_range_label = Make_Sliders_desired_values(app, "Range (kg): ", 0, 1, 7, math.ceil(max_range_row[10]/10)*10, (math.ceil(max_range_row[10]/10)*10)/2, 0, (math.ceil(max_range_row[10]/10)*10)/5)
    

    if max_range_row[10] == min_range_row[10]:
        Desired_range_slider, Desired_range_label = Make_Sliders_desired_values(app, "Range (km): ", math.ceil(max_range_row[10]), 1, 7, math.ceil(max_range_row[10])+1, math.ceil(max_range_row[10]), math.ceil(max_range_row[10])-1, 2)
    else:
        Desired_range_slider, Desired_range_label = Make_Sliders_desired_values(app, "Range (km): ", math.floor(min_range_row[10]), 1, 7, \
                                                                                    math.ceil(max_range_row[10]), \
                                                                                    ((math.ceil(max_range_row[10]) - (math.floor(min_range_row[10])))/2) + (math.floor(min_range_row[10])), \
                                                                                    (math.floor(min_range_row[10])), \
                                                                                    (math.ceil(max_range_row[10])) - (math.floor(min_range_row[10])))
    
    Desired_range_slider.configure(command=update_desired_range_label)
    Desired_range_slider.bind("<ButtonRelease-1>", on_desired_slider_release)
    # Desired_min_charging_time_slider, Desired_min_charging_time_label = Make_Sliders_desired_values(app, "Charging Time (10-80%) (mins): ", max_charging_time_row[11], 3, 7, (math.ceil(max_charging_time_row[11]/10)*10), (math.ceil(max_charging_time_row[11]/10)*10)/2, 0, (math.ceil(max_charging_time_row[11]/10)*10))
    # Desired_min_charging_time_slider.configure(command=update_desired_min_charging_time_label)
    # Desired_min_charging_time_slider.bind("<ButtonRelease-1>", on_desired_slider_release)
    # print(f"min list: {math.floor(min(list_km_per_min)*10)/10}")
    # print(f"max list: {math.ceil(max(list_km_per_min))}")

    # Desired_km_per_min_slider, Desired_km_per_min_label = Make_Sliders_desired_values(app, "Charging speed (km/min): ", (math.floor(min(list_km_per_min))), 3, 7, math.ceil(max(list_km_per_min)), math.ceil(max(list_km_per_min))/2, 0, math.ceil(max(list_km_per_min))*10)
    if min(list_km_per_min) == max(list_km_per_min):
        Desired_km_per_min_slider, Desired_km_per_min_label = Make_Sliders_desired_values(app, "Charging speed (km/min): ", math.ceil(max(list_km_per_min)), 3, 7, math.ceil(max(list_km_per_min))+1, math.ceil(max(list_km_per_min)), math.ceil(max(list_km_per_min))-1, 2)
    else:
        Desired_km_per_min_slider, Desired_km_per_min_label = Make_Sliders_desired_values(app, "Charging speed (km/min): ", math.floor(min(list_km_per_min)), 3, 7, \
                                                                                    math.ceil(max(list_km_per_min)), \
                                                                                    ((math.ceil(max(list_km_per_min)) - (math.floor(min(list_km_per_min))))/2) + (math.floor(min(list_km_per_min))), \
                                                                                    0, #(math.floor(min(list_km_per_min))), \
                                                                                    ((math.ceil(max(list_km_per_min))) - (math.floor(min(list_km_per_min))))*10)
        
    Desired_km_per_min_slider.configure(command=update_desired_km_per_min_label)
    Desired_km_per_min_slider.bind("<ButtonRelease-1>", on_desired_slider_release)
    # Desired_km_per_min_slider.set((math.floor(min(list_km_per_min)*10)/10))
    # desired_km_per_min = (0.8*max_range_row[10] - 0.1*max_range_row[10])/min_charging_time_row[11]

    if math.ceil((max_discharging_power_row[7]/1000)/10)*10 == math.ceil((min_discharging_power_row[7]/1000)/10)*10:
        Desired_max_discharge_power_slider, Desired_max_discharge_power_label = Make_Sliders_desired_values(app, "Max Discharge Power (kW): ", (min_discharging_power_row[7]/1000), 5, 7, (max_discharging_power_row[7]/1000)+1, (max_discharging_power_row[7]/1000), (max_discharging_power_row[7]/1000)-1, 2)
    else:
        Desired_max_discharge_power_slider, Desired_max_discharge_power_label = Make_Sliders_desired_values(app, "Max Discharge Power (kW): ", (min_discharging_power_row[7]/1000), 5, 7,\
                                                                                                            math.ceil((max_discharging_power_row[7]/1000)/10)*10, \
                                                                                                            (math.ceil((max_discharging_power_row[7]/1000)/10)*10 - (math.floor((min_discharging_power_row[7]/1000)/10)*10)/2) + (math.floor(((min_discharging_power_row[7])/1000)/10)*10), \
                                                                                                            math.floor(((min_discharging_power_row[7])/1000)/10)*10, \
                                                                                                            (math.ceil((max_discharging_power_row[7]/1000)/10)*10)-(math.floor((min_discharging_power_row[7]/1000)/10)*10)/10)
    Desired_max_discharge_power_slider.configure(command=update_desired_max_discharge_power_label)
    Desired_max_discharge_power_slider.bind("<ButtonRelease-1>", on_desired_slider_release)

    if max_mass_row[8] == min_mass_row[8]:
        Desired_max_mass_slider, Desired_max_mass_label = Make_Sliders_desired_values(app, "Max Mass (kg): ", math.ceil(max_mass_row[8]), 7, 7, math.ceil(max_mass_row[8])+1, math.ceil(max_mass_row[8]), math.ceil(max_mass_row[8])-1, 2)
    else:
        Desired_max_mass_slider, Desired_max_mass_label = Make_Sliders_desired_values(app, "Max Mass (kg): ", math.ceil(max_mass_row[8]), 7, 7, \
                                                                                    math.ceil(max_mass_row[8]), \
                                                                                    ((math.ceil(max_mass_row[8]) - (math.ceil(min_mass_row[8])))/2) + (math.ceil(min_mass_row[8])), \
                                                                                    (math.ceil(min_mass_row[8])), \
                                                                                    (math.ceil(max_mass_row[8])) - (math.ceil(min_mass_row[8])))
    Desired_max_mass_slider.configure(command=update_desired_max_mass_label)
    Desired_max_mass_slider.bind("<ButtonRelease-1>", on_desired_slider_release)
    
    on_desired_slider_release()

def excel_output():

    desired_values = [desired_range, desired_km_per_min, desired_max_discharging_power, desired_max_mass]
    matching_rows = [row for row in successful_combinations if row[10] >= desired_values[0] and (0.8*row[10] - 0.1*row[10])/row[11] >= desired_values[1] and row[7]/1000 >= desired_max_discharging_power and row[8] <= desired_max_mass]
    desired_EV_characteristics = [0, total_energy.get(), Pack_mass.get(), Max_V.get(), Min_V.get(), Discharging_power.get(), Charging_power.get()]
    slider_values = [0, float(total_energy_slider.get()), float(Pack_mass_slider.get()), float(Max_V_slider.get()), float(Min_V_slider.get()),float(Discharging_power_slider.get()), float(Charging_power_slider.get())]

    for i in range(len(desired_EV_characteristics)):
        # print(desired_EV_characteristics[i])

        if desired_EV_characteristics[i] == "":
            desired_EV_characteristics[i] = slider_values[i]
        elif i == 1: 
            try: desired_EV_characteristics[i] = float(total_energy.get())
            except: print("non number entered in energy")
        elif i == 2: 
            try: desired_EV_characteristics[i] = float(Pack_mass.get())
            except: print("non number entered in pack mass")
        elif i == 3: 
            try: desired_EV_characteristics[i] = float(Max_V.get())
            except: print("non number entered in max voltage")
        elif i == 4: 
            try: desired_EV_characteristics[i] = float(Min_V.get())
            except: print("non number entered in min voltage")
        elif i == 5: 
            try: desired_EV_characteristics[i] = float(Discharging_power.get())
            except: print("non number entered in discharging")
        elif i == 6: 
            try: desired_EV_characteristics[i] = float(Charging_power.get())
            except: 
                print("non number entered in charging")

     
    if len(matching_rows) == 1:
        
        data = {'Range': [matching_rows[0][10]], 'Charging time(10-80%)': [matching_rows[0][11]], 'Charging speed (km/min)': (0.8*matching_rows[0][10] - 0.1*matching_rows[0][10])/matching_rows[0][11], \
                'Max discharging power': [matching_rows[0][7]], 'Min pack mass': [matching_rows[0][8]],\
                'Max charging power': [matching_rows[0][9]], 'Actual Energy': [matching_rows[0][6]], 'Battery 1': [matching_rows[0][0]], 'Series Bat 1': [matching_rows[0][1]], 'Parallel bat 1': [matching_rows[0][2]], \
                'Battery 2': [matching_rows[0][3]], 'Series bat 2': [matching_rows[0][4]], 'Parallel bat 2': [matching_rows[0][5]], 'EV mass without pack': [car_data[0]], \
                'EV drag coefficient': [car_data[1]], 'EV frontal area': [car_data[2]], 'EV rolling resistance': [car_data[3]], \
                'Total energy': [desired_EV_characteristics[1]], 'Pack mass': [desired_EV_characteristics[2]], 'Max V': [desired_EV_characteristics[3]], 'Min V': [desired_EV_characteristics[4]], \
                'Discharging power': [desired_EV_characteristics[5]], 'Charging power': [desired_EV_characteristics[6]]}
        df = pd.DataFrame(data)

        file_path = 'Outputted Data.xlsx'
        sheet_name = 'Outputted Data'

        # Load existing workbook
        try:
            existing_book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Position the new data at the bottom of the existing sheet
                start_row = existing_book[sheet_name].max_row
                df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=1, index=False, header=False)
        except FileNotFoundError:
            # If file doesn't exist, create a new file
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print("Printed to file")
        
    else:
        print("Please select a single option to output")
    
    # existing_book.save(file_path)



def excel_output_all():

    desired_values = [desired_range, desired_km_per_min, desired_max_discharging_power, desired_max_mass]
    matching_rows = [row for row in successful_combinations if row[10] >= desired_values[0] and (0.8*row[10] - 0.1*row[10])/row[11] >= desired_values[1] and row[7]/1000 >= desired_max_discharging_power and row[8] <= desired_max_mass]

    desired_EV_characteristics = [0, total_energy.get(), Pack_mass.get(), Max_V.get(), Min_V.get(), Discharging_power.get(), Charging_power.get()]
    slider_values = [0, float(total_energy_slider.get()), float(Pack_mass_slider.get()), float(Max_V_slider.get()), float(Min_V_slider.get()),float(Discharging_power_slider.get()), float(Charging_power_slider.get())]

    for i in range(len(desired_EV_characteristics)):
        # print(desired_EV_characteristics[i])

        if desired_EV_characteristics[i] == "":
            desired_EV_characteristics[i] = slider_values[i]
        elif i == 1: 
            try: desired_EV_characteristics[i] = float(total_energy.get())
            except: print("non number entered in energy")
        elif i == 2: 
            try: desired_EV_characteristics[i] = float(Pack_mass.get())
            except: print("non number entered in pack mass")
        elif i == 3: 
            try: desired_EV_characteristics[i] = float(Max_V.get())
            except: print("non number entered in max voltage")
        elif i == 4: 
            try: desired_EV_characteristics[i] = float(Min_V.get())
            except: print("non number entered in min voltage")
        elif i == 5: 
            try: desired_EV_characteristics[i] = float(Discharging_power.get())
            except: print("non number entered in discharging")
        elif i == 6: 
            try: desired_EV_characteristics[i] = float(Charging_power.get())
            except: 
                print("non number entered in charging")

     

    full_data = []
    
    for i in range(len(matching_rows)):
        
        data = {'Range (km)': matching_rows[i][10], 'Charging time (10-80%)': matching_rows[i][11], 'Charging speed (km/min)': (0.8*matching_rows[i][10] - 0.1*matching_rows[i][10])/matching_rows[i][11],\
                'Max discharging power (kW)': (matching_rows[i][7]/1000), 'Min pack mass (kg)': matching_rows[i][8], 'Max charging power (kW)': (matching_rows[i][9]/1000),\
                'Pack energy (kWh)': (matching_rows[i][6]/1000), 'Battery 1': matching_rows[i][0], 'Series 1': matching_rows[i][1], 'Parallel 1': matching_rows[i][2], \
                'Battery 2': matching_rows[i][3], 'Series 2': matching_rows[i][4], 'Parallel 2': matching_rows[i][5], 'EV mass without pack (kg)': car_data[0], \
                'EV drag coefficient': car_data[1], 'EV frontal area (m²)': car_data[2], 'EV rolling resistance (N)': car_data[3], \
                'Required pack energy (kWh)': (desired_EV_characteristics[1]/1000), 'Pack mass (kg)': desired_EV_characteristics[2], 'Max voltage (V)': desired_EV_characteristics[3], 'Min voltage (V)': desired_EV_characteristics[4], \
                'Required discharging power (kW)': (desired_EV_characteristics[5]/1000), 'Required charging power (kW)': (desired_EV_characteristics[6]/1000)}
        
        full_data.append(data)

    # columns = ['Range (km)', 'Charging time(10-80%)', 'Charging speed (km/min)', 'Max discharging power', 'Min pack mass',
    #        'Max charging power', 'Actual Energy', 'Battery 1', 'Series Bat 1', 'Parallel bat 1',
    #        'Battery 2', 'Series bat 2', 'Parallel bat 2', 'EV mass without pack',
    #        'EV drag coefficient', 'EV frontal area', 'EV rolling resistance',
    #        'Total energy', 'Pack mass', 'Max V', 'Min V',
    #        'Discharging power', 'Charging power']
    
    df = pd.DataFrame(full_data)#, columns=columns)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    file_path = os.path.join(f"Outputs", f"Output_{timestamp}.xlsx")
    # sheet_name = 'Outputted Data'
    df.to_excel(file_path, index=False)



    # # Load existing workbook
    # try:
    #     existing_book = load_workbook(file_path)
    #     with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    #         # Position the new data at the bottom of the existing sheet
    #         start_row = existing_book[sheet_name].max_row
    #         df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=1, index=False, header=False)
    # except FileNotFoundError:
    #     # If file doesn't exist, create a new file
    #     with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
    #         df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print("Printed to file")
        
   
    
    # existing_book.save(file_path)
    

def BMS_option():

    battery_data, WLTP_data = Get_battery_and_WLTP_data()
    
    desired_values = [desired_range, desired_km_per_min, desired_max_discharging_power, desired_max_mass]
    matching_rows = [row for row in successful_combinations if row[10] >= desired_values[0] and (0.8*row[10] - 0.1*row[10])/row[11] >= desired_values[1] and row[7]/1000 >= desired_max_discharging_power and row[8] <= desired_max_mass]

    desired_EV_characteristics = [0, total_energy.get(), Pack_mass.get(), Max_V.get(), Min_V.get(), Discharging_power.get(), Charging_power.get()]
    slider_values = [0, float(total_energy_slider.get()), float(Pack_mass_slider.get()), float(Max_V_slider.get()), float(Min_V_slider.get()),float(Discharging_power_slider.get()), float(Charging_power_slider.get())]

    for i in range(len(desired_EV_characteristics)):
        # print(desired_EV_characteristics[i])

        if desired_EV_characteristics[i] == "":
            desired_EV_characteristics[i] = slider_values[i]
        elif i == 1: 
            try: desired_EV_characteristics[i] = float(total_energy.get())
            except: print("non number entered in energy")
        elif i == 2: 
            try: desired_EV_characteristics[i] = float(Pack_mass.get())
            except: print("non number entered in pack mass")
        elif i == 3: 
            try: desired_EV_characteristics[i] = float(Max_V.get())
            except: print("non number entered in max voltage")
        elif i == 4: 
            try: desired_EV_characteristics[i] = float(Min_V.get())
            except: print("non number entered in min voltage")
        elif i == 5: 
            try: desired_EV_characteristics[i] = float(Discharging_power.get())
            except: print("non number entered in discharging")
        elif i == 6: 
            try: desired_EV_characteristics[i] = float(Charging_power.get())
            except: 
                print("non number entered in charging")

     
    if len(matching_rows) == 1:

        # Gemini_possible = 0
        # PHEV_possible = 0
        # Both_possible = 0
        # HEV_possible = 0
        
        battery_data_series_parallel = [matching_rows[0][1], matching_rows[0][2], matching_rows[0][4], matching_rows[0][5]]
        battery_data_series_parallel_battery_1 = [matching_rows[0][1], matching_rows[0][2], 0, 0]
        battery_data_series_parallel_battery_2 = [matching_rows[0][4], matching_rows[0][5], 0, 0]

        battery_1 = battery_data[f"battery_{matching_rows[0][0]}_index"]
        battery_2 = battery_data[f"battery_{matching_rows[0][3]}_index"]

        # print(f"matching_rows: {matching_rows}")
        # print(f"battery_data_series_parallel: {battery_data_series_parallel}")
        # print(f"battery_1: {battery_1}")
        # print(f"battery_2: {battery_2}")

        Range_battery_1, Range_battery_2 = Range_Estimation_for_Each_Battery(WLTP_data, car_data, battery_data_series_parallel, battery_1, battery_2)
        total_range = Range_battery_1 + Range_battery_2
        # print(f"Range 1: {Range_battery_1}, Range 2: {Range_battery_2}")
        Power_battery_1, Power_battery_2 = Power_Estimation_for_Each_Battery(battery_data_series_parallel, battery_1, battery_2)
        # print(f"Power 1: {Power_battery_1}, Power 2: {Power_battery_2}")

        
        if battery_1[10] == 'not defined':
            battery_1_cycles = 1750
        else:
            battery_1_cycles = battery_1[10]
        if battery_2[10] == 'not defined':
            battery_2_cycles = 1750
        else:
            battery_2_cycles = battery_2[10]

        try: req_power = float(Discharging_power.get())
        except: req_power = float(Discharging_power_slider.get())

        Gemini_1 = 0
        Gemini_2 = 0
        PHEV_1 = 0
        PHEV_2 = 0
        HEV_1 = 0
        HEV_2 = 0
        Both = 0


        # Check options
        if (battery_1_cycles < 400 and Range_battery_2 > 82.86 and Power_battery_2 > req_power and battery_2_cycles > 1500):
            Gemini_1 = 1 # Gemini has to be used to protect the battery
        if (battery_2_cycles < 400 and Range_battery_1 > 82.86 and Power_battery_1 > req_power and battery_1_cycles > 1500):
            Gemini_2 = 1

        
        # if (Range_battery_1 < 8 and total_range > 150 and Power_battery_2 < req_power):
        #         HEV_1 = 1 # Have to use HEV because the range is so low on one battery that it can only be used as a power boost
        # if (Range_battery_2 < 8 and total_range > 150 and Power_battery_1 < req_power):
        #         HEV_2 = 1
        if (Range_battery_1 < Range_battery_2 and Power_battery_2 < req_power and battery_1_cycles > 2500):
                HEV_1 = 1 # Have to use HEV because the range is so low on one battery that it can only be used as a power boost
        if (Range_battery_2 < Range_battery_1 and Power_battery_1 < req_power and battery_2_cycles > 2500):
                HEV_2 = 1
        
        if (Power_battery_1 > req_power and Range_battery_1 > 48.28 and Range_battery_2 > Range_battery_1 and battery_1_cycles > 2000):
            PHEV_1 = 1
        if (Power_battery_2 > req_power and Range_battery_2 > 48.28 and Range_battery_2 < Range_battery_1 and battery_2_cycles > 2000):
            PHEV_2 = 1

        if (Gemini_1 + Gemini_2 + PHEV_1 + PHEV_2 + HEV_1 + HEV_2) == 0:
            Both = 1 
            New_Window()
            BMS_label.configure(text=f"The Dual System is the best option to use with these batteries\n")
            print("Both possible")
        elif(Gemini_1 + Gemini_2 + PHEV_1 + PHEV_2 + HEV_1 + HEV_2) > 1:
            print("More than one option selected")
        
        else:
            if Gemini_1 == 1:
                New_Window()
                BMS_label.configure(text=f"The Gemini BMS option should be used with battery 1 as the extended range battery\n"
                                         f"and battery 2 as the everyday use battery.\n"
                                         f"The range of everyday mode is: {Range_battery_2:.1f} km\n"
                                         f"The range of the extending battery is: {Range_battery_1:.1f} km\n"
                                         f"Total range: {total_range:.1f} km\n"
                                         f"The power of everyday mode is: {Power_battery_2/1000} kW\n"
                                         f"The cycle life of battery 1: {battery_1_cycles}\n")
                
                print("Gemini 1 possible")
            if Gemini_2 == 1:   
                New_Window()
                BMS_label.configure(text=f"The Gemini BMS option should be used with battery 2 as the extended range battery\n"
                                         f"and battery 1 as the everyday use battery.\n"
                                         f"The range of everyday mode is: {Range_battery_1:.1f} km\n"
                                         f"The range of the extending battey is: {Range_battery_2:.1f} km\n"
                                         f"Total range: {total_range:.1f} km\n"
                                         f"The power of everyday mode is: {Power_battery_1/1000} kW\n"
                                         f"The cycle life of battery 2: {battery_2_cycles}\n")
                print("Gemini 2 possible")
            if PHEV_1 == 1:
                New_Window()
                BMS_label.configure(text=f"The PHEV BMS option should be used with battery 2 as the extended range battery\n"
                                         f"and battery 1 as the everyday use battery.\n"
                                         f"The range of everyday mode is: {Range_battery_1:.1f} km\n"
                                         f"The range of the extending battery is: {Range_battery_2:.1f} km\n"
                                         f"Total range: {total_range:.1f} km\n"
                                         f"The power of everyday mode is: {Power_battery_1/1000} kW\n")
                print("PHEV 1 possible")
            if PHEV_2 == 1:
                New_Window()
                BMS_label.configure(text=f"The PHEV BMS option should be used with battery 1 as the extended range battery\n"
                                         f"and battery 2 as the everyday use battery.\n"
                                         f"The range of everyday mode is: {Range_battery_2:.1f} km\n"
                                         f"The range of the extending battery is: {Range_battery_1:.1f} km\n"
                                         f"Total range: {total_range:.1f} km\n"
                                         f"The power of everyday mode is: {Power_battery_2/1000} kW\n")  
                print("PHEV 2 possible")
            if HEV_1 == 1:
                New_Window()
                BMS_label.configure(text=f"The HEV BMS option should be used with battery 2 as the main battery\n"
                                         f"and battery 1 used as the extra power source for acceleration.\n"
                                         f"The power of the main battery is: {Power_battery_2/1000} kW\n"
                                         f"The power of the suplimentary power source battery is: {Power_battery_1/1000} kW\n"
                                         f"The total power of the battery is: {(Power_battery_1+Power_battery_2)/1000} kW\n"
                                         f"Total range: {total_range:.1f} km\n")
                print("HEV 1 possible")
            if HEV_2 == 1:
                New_Window()
                BMS_label.configure(text=f"The HEV BMS option should be used with battery 1 as the main battery\n"
                                         f"and battery 2 used as the extra power source for acceleration.\n"
                                         f"The power of the main battery is: {Power_battery_1/1000} kW\n"
                                         f"The power of the suplimentary power source battery is: {Power_battery_2/1000} kW\n"
                                         f"The total power of the battery is: {(Power_battery_1+Power_battery_2)/1000} kW\n"
                                         f"Total range: {total_range:.1f} km\n")
                print("HEV 2 possible")
        
        print(f"Range 1: {Range_battery_1:.1f}, Range 2: {Range_battery_2:.1f}")
        print(f"Power 1: {Power_battery_1:.1f}, Power 2: {Power_battery_2:.1f}")
        print(f"Cycles 1: {battery_1_cycles}, Cycles 2: {battery_2_cycles}")

        # else: 
        #     Both_possible = 0
        #     HEV_possible = 0
        #     Gemini_possible = 0
        #     PHEV_possible = 0
        
        #     if Power_battery_1 < float(Discharging_power.get()) and Power_battery_2 < float(Discharging_power.get()):
        #         Both_possible += 1
        #         HEV_possible += 1
        #         # option = [Gemini_possible, PHEV_possible, Both_possible, HEV_possible]
            
        #     elif Range_battery_1 < (total_range*0.65) and Range_battery_2 < (total_range*0.65):
        #         Both_possible += 1
        #         PHEV_possible += 1
        #         # option = [Gemini_possible, PHEV_possible, Both_possible, HEV_possible]
            

        #     elif (Range_battery_1 > Range_battery_2 and Power_battery_1 < float(Discharging_power.get()) and Power_battery_2 > float(Discharging_power.get())) or (Range_battery_2 > Range_battery_1 and Power_battery_2 < float(Discharging_power.get()) and Power_battery_2 > float(Discharging_power.get())):
        #         if Range_battery_1 > Range_battery_2 and 

        # power of energy dense one isn't enough to power the car
    
    else:
        print("Please select a single option to output")
      
def New_Window():
    global BMS_window, BMS_label

    BMS_window = ctk.CTkToplevel()
    BMS_window.title("Which BMS option should be used")
    BMS_window.geometry("450x150")
    BMS_label = ctk.CTkLabel(BMS_window, text="", anchor="w", justify='left')
    BMS_label.grid(padx = 10, pady = 10, row=0, column=0, sticky="w")
    

# Total Energy
total_energy = ctk.CTkEntry(app, placeholder_text="Total Energy", width=160, height=28)
total_energy.grid(row= 3, column= 1, padx=10, pady=10)
total_energy_slider = ctk.CTkSlider(app, from_= 0, to=150000, number_of_steps=150)
total_energy_slider.grid(row= 3, column=2, padx=10, pady=10)
total_energy_slider.set(input_energy)
# total_energy_slider.configure(command=update_total_energy_label)
total_energy_label = ctk.CTkLabel(app, text=input_energy)
total_energy_label.grid(row= 4, column=2)

# Desired EV characteristics
# EV_range, EV_range_slider, EV_range_label = Make_Entry_boxes_and_sliders(app, f"Range: ", input_range, 1, 2, 1000, 500, 0, 200, "Range (km)")
total_energy, total_energy_slider, total_energy_label = Make_Entry_boxes_and_sliders(app, f"Energy (Wh): ", input_energy, 1, 2, 150000, 75000, 0, 150, "Total Energy (Wh)")
Pack_mass, Pack_mass_slider, Pack_mass_label = Make_Entry_boxes_and_sliders(app, f"Pack Mass (kg): ", input_max_mass_pack, 3, 2, 1000, 500, 0, 200, "Pack Mass (kg)")
Max_V, Max_V_slider, Max_V_label = Make_Entry_boxes_and_sliders(app, f"Maximum Voltage (V): ", input_max_V, 5, 2, 1000, 500, 0, 200, "Maximum Voltage (V)")
Min_V, Min_V_slider, Min_V_label = Make_Entry_boxes_and_sliders(app, f"Minimum Voltage (V): ", input_min_V, 7, 2, 600, 300, 0, 120, "Minimum Voltage (V)")
Discharging_power, Discharging_power_slider, Discharging_power_label = Make_Entry_boxes_and_sliders(app, f"Discharging Power (W): ", input_discharging_power, 9, 2, 500000, 250000, 0, 500, "Peak Discharging Power (W)")
Charging_power, Charging_power_slider, Charging_power_label = Make_Entry_boxes_and_sliders(app, f"Charging Power (W): ", input_charging_power, 11, 2, 300000, 150000, 0, 300, "Peak Charging Power (W)")
Volume, Volume_slider, Volume_label = Make_Entry_boxes_and_sliders(app, f"Pack Volume (m³): ", input_max_volume, 13, 2, 2, 1, 0, 400, "Pack Volume (m³)")

# EV_range_slider.configure(command=update_range_label)
total_energy_slider.configure(command=update_total_energy_label)
Pack_mass_slider.configure(command=update_Pack_mass_label)
Max_V_slider.configure(command=update_Max_V_label)
Min_V_slider.configure(command=update_Min_V_label)
Discharging_power_slider.configure(command=update_Discharging_power_label)
Charging_power_slider.configure(command=update_Charging_power_label)
Volume_slider.configure(command=update_Volume_label)


# EV metrics

# Needed to stop the sliders from going off the ends
EV_mass = ctk.CTkEntry(app, placeholder_text="EV Mass Without Pack (kg)", width=160, height=28)
EV_mass.grid(row= 1, column= 3, padx=10, pady=10)
EV_mass_slider = ctk.CTkSlider(app, from_= 0, to=5000, number_of_steps=2500)
EV_mass_slider.grid(row= 1, column=4, padx=10, pady=10)
EV_mass_slider.set(car_data[0])
# total_energy_slider.configure(command=update_total_energy_label)
EV_mass_label = ctk.CTkLabel(app, text=car_data[0])
EV_mass_label.grid(row= 2, column=4)

# EV mass without battery pack
EV_mass, EV_mass_slider, EV_mass_label = Make_Entry_boxes_and_sliders(app, f"EV Mass Without Pack (kg): ", car_data[0], 1, 4, 5000, 2500, 0, 1000, "EV Mass Without Pack (kg)")
EV_mass_slider.configure(command=update_EV_mass_label)
# EV drag coefficient
EV_drag, EV_drag_slider, EV_drag_label = Make_Entry_boxes_and_sliders(app, f"EV Drag Coefficient: ", car_data[1], 3, 4, 1, 0.5, 0, 100, "EV Drag Coefficient")
EV_drag_slider.configure(command=update_EV_drag_label)
# EV frontal area
EV_front_area, EV_front_area_slider, EV_front_area_label = Make_Entry_boxes_and_sliders(app, f"EV Frontal Area (m²): ", car_data[2], 5, 4, 5, 2.5, 0, 200, "EV Frontal Area (m²)")
EV_front_area_slider.configure(command=update_EV_front_area_label)
# EV rolling resistance
EV_r_r, EV_r_r_slider, EV_r_r_label = Make_Entry_boxes_and_sliders(app, f"EV Rolling Resistance (N): ", car_data[3], 7, 4, 0.04, 0.015, 0, 40, "EV Rolling Resistance (N)")
EV_r_r_slider.configure(command=update_EV_r_r_label)

def Make_weightings_sliders():

    global Range_weighting_slider, Range_weighting_label, Min_charging_time_weighting_slider, Min_charging_time_weighting_label, max_discharge_power_weighting_slider, max_discharge_power_weighting_label, min_pack_mass_weighting_slider, min_pack_mass_weighting_label, min_charge_power_weighting_slider, min_charge_power_weighting_label
    # default_weighting = [1, 1, 1, 1, 1]

    # Needed to stop the sliders from going off the ends
    # EV_mass = ctk.CTkEntry(app, placeholder_text="EV Mass (kg)", width=160, height=28)
    # EV_mass.grid(row= 1, column= 3, padx=10, pady=10)
    Range_weighting_slider = ctk.CTkSlider(app, from_= 0, to=3, number_of_steps=300)
    Range_weighting_slider.grid(row= 1, column=6, padx=10, pady=10)
    Range_weighting_slider.set(car_data[0])
    # total_energy_slider.configure(command=update_total_energy_label)
    Range_weighting_label = ctk.CTkLabel(app, text=default_weighting[0])
    Range_weighting_label.grid(row= 2, column=6)

    Range_weighting_slider, Range_weighting_label = Make_Sliders(app, "Range Weighting: ", default_weighting[0], 1, 6, 3, 1.5, 0, 60)
    Range_weighting_slider.configure(command=update_range_weighting_label)
    Range_weighting_slider.bind("<ButtonRelease-1>", on_weighting_slider_release)

    Min_charging_time_weighting_slider, Min_charging_time_weighting_label = Make_Sliders(app, "Max Charging Speed Weighting: ", default_weighting[1], 3, 6, 3, 1.5, 0, 60)
    Min_charging_time_weighting_slider.configure(command=update_min_charging_time_weighting_label)
    Min_charging_time_weighting_slider.bind("<ButtonRelease-1>", on_weighting_slider_release)

    max_discharge_power_weighting_slider, max_discharge_power_weighting_label = Make_Sliders(app, "Peak Discharge Power Weighting: ", default_weighting[2], 5, 6, 3, 1.5, 0, 60)
    max_discharge_power_weighting_slider.configure(command=update_max_discharge_power_weighting_label)
    max_discharge_power_weighting_slider.bind("<ButtonRelease-1>", on_weighting_slider_release)

    min_pack_mass_weighting_slider, min_pack_mass_weighting_label = Make_Sliders(app, "Min Pack Mass Weighting: ", default_weighting[3], 7, 6, 3, 1.5, 0, 60)
    min_pack_mass_weighting_slider.configure(command=update_min_pack_mass_weighting_label)
    min_pack_mass_weighting_slider.bind("<ButtonRelease-1>", on_weighting_slider_release)


# # min_charge_power_weighting_slider, min_charge_power_weighting_label = Make_Sliders(app, "Min Charge Power Weighting: ", default_weighting[4], 9, 6, 3, 1.5, 0, 60)
# # min_charge_power_weighting_slider.configure(command=update_min_charge_power_weighting_label)
# # min_charge_power_weighting_slider.bind("<ButtonRelease-1>", on_weighiting_slider_release)

# Desired_range_slider, Desired_range_label = Make_Sliders_desired_values(app, "Range (km): ", 350, 1, 7, 500, 2500, 0, 50)
# Desired_range_slider.configure(command=update_desired_range_label)
# Desired_range_slider.bind("<ButtonRelease-1>", on_desired_slider_release)

# Desired_min_charging_time_slider, Desired_min_charging_time_label = Make_Sliders_desired_values(app, "Charging Time(10-80%) (mins): ", 20, 3, 7, 90, 45, 0, 90)
# Desired_min_charging_time_slider.configure(command=update_desired_min_charging_time_label)
# Desired_min_charging_time_slider.bind("<ButtonRelease-1>", on_desired_slider_release)


    


# EV_range.bind("<KeyRelease>", update_sliders)
total_energy.bind("<KeyRelease>", update_sliders)  
Pack_mass.bind("<KeyRelease>", update_sliders)
Max_V.bind("<KeyRelease>", update_sliders)  
Min_V.bind("<KeyRelease>", update_sliders)
Discharging_power.bind("<KeyRelease>", update_sliders)  
Charging_power.bind("<KeyRelease>", update_sliders)
Volume.bind("<KeyRelease>", update_sliders)
EV_mass.bind("<KeyRelease>", update_sliders)
EV_drag.bind("<KeyRelease>", update_sliders)
EV_front_area.bind("<KeyRelease>", update_sliders)
EV_r_r.bind("<KeyRelease>", update_sliders)


# Button to calculate the result
calc_button = ctk.CTkButton(app, text="Calculate", command=calculate)
calc_button.grid(row= 10, column= 4, padx=10, pady=0)
# Scatter plot button
plot_button = ctk.CTkButton(app, text="All Options Scatter Plot", command=lambda: plot_scatter(successful_combinations_1_bat, successful_combinations_2_bat))
plot_button.grid(row=14, column=8, pady=10, padx=0)
plot_button.grid_forget()

plot_button_current_options = ctk.CTkButton(app, text="Current Options Scatter Plot", command=lambda: plot_scatter_current_options(successful_combinations_1_bat, successful_combinations_2_bat, \
                                                                                                            desired_range, desired_km_per_min, desired_max_discharging_power, desired_max_mass))
plot_button_current_options.grid(row=15, column=8, pady=0, padx=0)
plot_button_current_options.grid_forget()

parallel_coordinates_plot_button = ctk.CTkButton(app, text="Parallel Coordinates Plot", command=lambda: create_parallel_coordinates_plot(successful_combinations, has_calculate_been_pressed))
parallel_coordinates_plot_button.grid(row=15, column=8, pady=0, padx=0)
parallel_coordinates_plot_button.grid_forget()

# Button to reset the weightings
reset_weightings_button = ctk.CTkButton(app, text="Reset Weightings", command=reset_weightings)
reset_weightings_button.grid(row= 10, column= 6, padx=0, pady=0)
reset_weightings_button.grid_forget()

# Button to reset the inputs
reset_inputs_button = ctk.CTkButton(app, text="Reset Inputs", command=reset_inputs)
reset_inputs_button.grid(row= 10, column= 3, padx=0, pady=0)

# Button to output data to excel
excel_output_button = ctk.CTkButton(app, text="Excel Output 1", command=excel_output)
excel_output_button.grid(row= 9, column= 8, padx=0, pady=0)
excel_output_button.grid_forget()

excel_output_all_button = ctk.CTkButton(app, text="Excel Output", command=excel_output_all)
excel_output_all_button.grid(row= 10, column= 8, padx=0, pady=0)
excel_output_all_button.grid_forget()

# Button to determine which BMS to use
BMS_option_button = ctk.CTkButton(app, text="Determine BMS Option", command=BMS_option)
BMS_option_button.grid(row= 11, column= 8, padx=0, pady=0)
BMS_option_button.grid_forget()




# Create a label to display the result
result_label_weightings = ctk.CTkLabel(app, text=" \n \n \n \n \n ")
result_label_weightings.grid(row= 11, column= 6, padx=0, pady=5, rowspan=3) 

# Create a label to display the result of desired sliders
result_label_desired_values = ctk.CTkLabel(app, text="")
result_label_desired_values.grid(row= 11, column= 7, padx=0, pady=5, rowspan=3, sticky="w")

calculate_button_label = ctk.CTkLabel(app, text="Press calculate to determine\npotential battery combinations")
calculate_button_label.grid(row= 11, column= 4, padx=0, pady=5, rowspan = 2)

desired_EV_label = ctk.CTkLabel(app, text="Please enter desired EV chracteristics values or adjust slider:\n ")
desired_EV_label.grid(row=0, column= 1, columnspan=2, padx=20, pady=10)

EV_metrics_label = ctk.CTkLabel(app, text="Please enter EV metrics values or adjust slider:\n")
EV_metrics_label.grid(row=0, column= 3, columnspan=2, padx=20, pady=10)

desired_EV_outputs_title_label = ctk.CTkLabel(app, text="                        ")
desired_EV_outputs_title_label.grid(row=0, column= 7, columnspan=1, padx=5, pady=10)

Weightings_sliders_title_label = ctk.CTkLabel(app, text="                        ")
Weightings_sliders_title_label.grid(row=0, column= 6, columnspan=1, padx=5, pady=10)

# excel_output_button_label = ctk.CTkLabel(app, text="Press to output data to excel")
# excel_output_button_label.grid(row=13, column= 4, columnspan=1, padx=5, pady=10)






# checkbox_Boxes_to_sliders = ctk.CTkCheckBox(app, text="Transfer numbers entered to sliders?", command=update_sliders, variable=ctk.IntVar(value=1))
# checkbox_Boxes_to_sliders.grid(row=22, column=1, padx=(5, 5), pady=(0, 0), sticky="w")
# checkbox_disclamer_label = ctk.CTkLabel(app, text="If un-checked the program will\nuse the box values as default  ")
# checkbox_disclamer_label.grid(row= 23, column= 1, padx=10, pady=0)

# selected_option = ctk.IntVar(value=0)

checkbox_Only_2_batteries = ctk.CTkCheckBox(app, text="Only show 2 battery options", command=lambda: check_selected(1))#, variable=selected_option, onvalue=1, offvalue=0)
checkbox_Only_2_batteries.grid(row=15, column=7, padx=(5, 5), pady=(0, 0), sticky="w")
checkbox_Only_2_batteries.grid_forget()

checkbox_Only_1_battery = ctk.CTkCheckBox(app, text="Only show 1 battery options", command=lambda: check_selected(2))#, variable=selected_option, onvalue=1, offvalue=0)
checkbox_Only_1_battery.grid(row=16, column=7, padx=(5, 5), pady=(0, 0), sticky="w")

checkbox_Only_1_battery.grid_forget()

checkbox_both_batteries = ctk.CTkCheckBox(app, text="Show all battery options", command=lambda: check_selected(3), variable=ctk.IntVar(value=1))#, variable=selected_option, onvalue=1, offvalue=0)
checkbox_both_batteries.grid(row=17, column=7, padx=(5, 5), pady=(0, 0), sticky="w")
checkbox_both_batteries.grid_forget()


# Verticle lines
vertical_line = ctk.CTkFrame(app, width=2, height=475, fg_color="gray")
vertical_line.place(x=832, y=0)
vertical_line.grid_forget()


# Start the main event loop
app.mainloop()

